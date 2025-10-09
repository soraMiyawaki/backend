# app/routers/attendance.py
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, time
import openpyxl
from pathlib import Path
import tempfile
import shutil

router = APIRouter(prefix="/api/attendance", tags=["attendance"])

# Template path - use local copy for better performance
TEMPLATE_PATH = Path(__file__).parent.parent.parent / "template.xlsx"
# Fallback to OneDrive path if local template doesn't exist
if not TEMPLATE_PATH.exists():
    TEMPLATE_PATH = Path(r"C:\Users\Exitotrinity-13\OneDrive\Desktop\ひな型\勤務管理表_2023_12 (4).xlsx")

class DailyAttendance(BaseModel):
    day: int  # 1-31
    start_time: Optional[str] = None  # "09:00"
    end_time: Optional[str] = None    # "18:00"
    is_holiday: bool = False

class AttendanceRequest(BaseModel):
    year: int
    month: int
    employee_name: str = ""
    employee_id: str = ""
    attendance_data: List[DailyAttendance]

def time_str_to_excel(time_str: str) -> str:
    """Convert time string '09:00' to Excel time format '09:00:00'"""
    if not time_str:
        return ""
    return f"{time_str}:00"

@router.post("/generate")
async def generate_attendance_excel(request: AttendanceRequest):
    """
    勤怠管理表Excelファイルを生成
    """
    print(f"[ATTENDANCE] Received request for {request.year}/{request.month}")
    print(f"[ATTENDANCE] Template path: {TEMPLATE_PATH}")
    print(f"[ATTENDANCE] Template exists: {TEMPLATE_PATH.exists()}")

    try:
        # Load template
        if not TEMPLATE_PATH.exists():
            error_msg = f"Template file not found at: {TEMPLATE_PATH}"
            print(f"[ATTENDANCE] ERROR: {error_msg}")
            raise HTTPException(status_code=500, detail=error_msg)

        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active

        # Set year and month
        ws['B1'] = request.year
        ws['E1'] = request.month

        # Set employee info (if provided)
        if request.employee_name:
            ws['G5'] = request.employee_name  # Assuming this cell for employee name
        if request.employee_id:
            ws['B5'] = request.employee_id    # Assuming this cell for employee ID

        # Fill attendance data
        for daily in request.attendance_data:
            if daily.day < 1 or daily.day > 31:
                continue

            row_index = 10 + daily.day  # Row 11 is day 1

            if daily.is_holiday:
                # Clear times for holidays
                ws[f'D{row_index}'] = ""
                ws[f'E{row_index}'] = ""
            else:
                # Set start/end times
                if daily.start_time:
                    ws[f'D{row_index}'] = time_str_to_excel(daily.start_time)
                if daily.end_time:
                    ws[f'E{row_index}'] = time_str_to_excel(daily.end_time)

        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        wb.close()

        # Prepare filename
        filename = f"勤務管理表_{request.year}_{request.month:02d}.xlsx"

        return FileResponse(
            tmp_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
